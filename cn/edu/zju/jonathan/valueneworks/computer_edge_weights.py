import matplotlib.pyplot as plt
import math
import os
import networkx as nx
import numpy as np
import pandas as pd
from random import *
import time
import logging
import xlsxwriter
from openpyxl import load_workbook
from xlsxwriter import Workbook

logger = logging.getLogger()
logger.setLevel(logging.INFO)
# 定义handler的输出格式
#logger to console
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s")
#fh.setFormatter(formatter)
ch.setFormatter(formatter)
logger.addHandler(ch)

q_start_columm=13
rca_threshold=2.1
#equations = "score1-5"
equations = "eq-nonlinear"


#get the mapped score
def getscore (x):
    b = 6
    c = 10
    return x**2 - b*x +c

#logger.info(getscore(5))
#exit(0)

f1 = 'D:\\py\\data\\jonathan\\MCS_recoded-q.csv'
f_code= 'D:\\py\\data\\jonathan\\code.xlsx'
f_rca= 'D:\\py\\data\\jonathan\\%s\\rca_number_%.1f.xlsx' % (equations, rca_threshold)
f_weights = 'D:\\py\\data\\jonathan\\%s\\weights_number_min_%.1f.xlsx' % (equations, rca_threshold)
logger.info("RCA:%.1f,Equation:%s" % (rca_threshold,equations))
if os.path.isfile(f_rca)==False:
    country_paticipant_quest_rca = pd.DataFrame(columns=["country", 'id', 'Q', 'RCA'])
    country_quest_phi_weight_network = pd.DataFrame(columns=["country", 'QA', 'QB', 'weights'])

    writer = pd.ExcelWriter(f_rca)
    country_paticipant_quest_rca.to_excel(writer,  index=False)
    writer.save()

    writer = pd.ExcelWriter(f_weights)
    country_quest_phi_weight_network.to_excel(writer, index=False)
    writer.save()

dirty_words = {'Not sure', 'Dont\'t know','Refused'}

logger.info("loading data...")
time_start=time.time()
questions=['q1', 'q2', 'q3', 'q4',
                 'q5', 'q6', 'q7', 'q8',
                 'q9', 'q10', 'q11', 'q12',
                 'q13', 'q14', 'q15', 'q16']
df = pd.read_csv(f1)
#df = pd.read_excel(f1)
#df = df.loc[::, ['id','country', 'q1', 'q2', 'q3', 'q4',
#                 'q5', 'q6', 'q7', 'q8',
#                 'q9', 'q10', 'q11', 'q12',
#                 'q13', 'q14', 'q15', 'q16']]
#df_code= pd.read_excel(f_code)
#df_code = df_code.loc[::, ['q', 'answer', 'coded_value']]

time_end = time.time()
logger.info("data loaded...time cost:%d s'", time_end - time_start)
countries = {}
data = df.values.tolist()
countries = [data[i][1] for i in range(len(data))]
countries = sorted(list(set(countries)))

#logger.info(data[0])
#logger.info(data[1])
#logger.info(data[2])
#exit(0)

#logger.info(countries)
#logger.info("total # of countries: %d", len(countries))
#logger.info(countries)
#exit(0)
#code = [data[i][3] for i in range(len(data))]

#embbing answers to numerical values
#answer_values ={}
#data_answers = df_code.values.tolist()
#for i in range(len(data_answers)):
#   q = data_answers[i][0]
#   answer= data_answers[i][1]
#   coded_value= data_answers[i][2]
#   if(answer_values.get(q)==None):
#       answer_values[q]={}
#   answer_values[q][answer] = coded_value

#logger.info(answer_values)
#exit(0)

#current_countries =countries
k=56


#current_countries = [countries[i] for i in range(k,len(countries))]
#current_countries =["Zambia"]
current_countries =["Germany","Hong Kong","United States","Zambia"]

#logger.info(current_countries)
#logger.info(countries)

#df_one_country=[]

for y in countries:
    if (y in current_countries):
        k += 1
        time_end = time.time()
        country_paticipant_quest_rca = pd.DataFrame(columns=["country", 'id', 'Q', 'RCA'])
        country_quest_phi_weight_network = pd.DataFrame(columns=["country", 'QA', 'QB', 'weights'])
        df_one_country = df.loc[df['country'] == y]
        logger.info('Country: %s(%d/%d),# of participants:%d, time cost:%d s',y, k,len(countries), len(df_one_country), time_end - time_start)
        data_one_country = df_one_country.values.tolist()
        df_one_country_values =  pd.DataFrame(columns=['id','q1', 'q2', 'q3', 'q4',
                 'q5', 'q6', 'q7', 'q8',
                 'q9', 'q10', 'q11', 'q12',
                 'q13', 'q14', 'q15', 'q16'])

        for i in range(len(data_one_country)):
            q_values = [data_one_country[i][0]] #id
            for j in range(len(questions)):
#              if data_one_country[i][j+2] in dirty_words:
#                    q_values.append(0)  #set defalut dirty words as zero
#              else:
                    #q_values.append(answer_values[questions[j]][data_one_country[i][j+2]])
                    v= getscore(data_one_country[i][j+q_start_columm])
                    q_values.append(v)
            df_one_country_values.loc[len(df_one_country_values)] = q_values
#        logger.info(df_one_country_values)
#        logger.info(len(df_one_country_values))
#        exit(0)
        #code = [data_one_country[i][2] + str(data_one_country[i][3]) for i in range(len(data_one_country))]
        #df_one_country.insert(loc=len(df_one_country.columns), column='行业代码', value=code)
        paticipants = [data_one_country[i][0] for i in range(len(df_one_country))]
        paticipants = sorted(list(set(paticipants))) #去重

        #计算每个问题的全局权重
        question_part={}
        df_temp = df_one_country_values.drop(['id'], axis=1)
#        logger.info(df_temp)
        sum_quesitons=df_temp.sum()
        sum_all= sum(sum_quesitons)
        sum_participants=df_temp.sum(axis=1)

#        logger.info(sum_all)
#        logger.info(sum_quesitons)
#        exit(0)
        question_participant = {}
        for i in range(len(questions)):
            q = questions[i]
            question_part[q] = sum_quesitons[i] / sum_all
            question_participant[q]=[]
            #logger.info(c,code_company.get(c))

        #logger.info(question_part)
        #exit(0)
        #计算单一question的RCA
        i=0
        for participant in paticipants:
            #ignore invalid users : did not response at all
            if (sum_participants[i] >0):
                for j in range(len(questions)):
#                logger.info('est')
#                logger.info(df_one_country_values.loc[(df_one_country_values["id"]==participant)])
                #logger.info(df_one_country_values.loc[(df_one_country_values["id"]==participant)][questions[j]])
#                logger.info(df_one_country_values.loc[i,questions[j]])
#                logger.info(sum_participants[i])
                    participant_question_ratio= df_one_country_values.loc[i,questions[j]] /sum_participants[i]
                    #participant_question_ratio= df_one_country_values.loc["id"==][questions[j]]/sum_participants[i]
                    #participant_question_global_ratio = df_one_country_values[i][j+1]/ df_one_country_values[i].sum(axis=0) #question_value
                    #binary computing RCA
    #                logger.info(participant_question_ratio)
    #                logger.info(question_part[questions[j]])
    #                exit(0)
    #                if df_one_country_values.loc[i,questions[j]]<=0:
                        #participant_question_ratio = 0
                    rca = 1 if participant_question_ratio/(question_part[questions[j]]) >= rca_threshold else 0
                    country_paticipant_quest_rca.loc[len(country_paticipant_quest_rca)] = [y, participant,questions[j], rca]

                   ###记录有效rca对应的人
                    if rca == 1:
                        list_participants=question_participant.get(questions[j])
    #                    logger.info(participant_question_ratio)
    #                    logger.info(question_part[questions[j]])
    #                    logger.info(participant)
    #                    exit(0)
                        list_participants.append(participant)
                        question_participant[questions[j]]=list_participants
            i += 1
        #计算phi
        for i in range(len(questions)-1):
            for j in range(i+1, len(questions)):
                #两个集合的交集
                set_participants= set(question_participant[questions[i]]) & set(question_participant[questions[j]])
                if (len(set_participants)>0):
                    phi_ij = len(set_participants) / len(question_participant[questions[j]])
                    phi_ji = len(set_participants) / len(question_participant[questions[i]])
                    #@Todo: check  min or max
                    phi = phi_ij if phi_ij < phi_ji else phi_ji
                    country_quest_phi_weight_network.loc[len(country_quest_phi_weight_network)] = [y,questions[i], questions[j],phi]
                #else:
                #    country_quest_phi_weight_network.loc[len(country_quest_phi_weight_network)] = [y,questions[i], questions[j], 0]

        #write to file

        #time_end = time.time()
        writer = pd.ExcelWriter(f_weights, engine='openpyxl',mode ='a', if_sheet_exists='overlay')
        book=load_workbook(f_weights)
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df1=pd.DataFrame(pd.read_excel(f_weights,sheet_name='Sheet1'))
        df_rows = df1.shape[0]
        country_quest_phi_weight_network.to_excel(writer,sheet_name='Sheet1',startrow=df_rows+1,index=False,header=False)
        writer.save()
        writer.close()

        writer2 = pd.ExcelWriter(f_rca, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        book = load_workbook(f_rca)
        writer2.book = book
        writer2.sheets = dict((ws.title, ws) for ws in book.worksheets)
        df2 = pd.DataFrame(pd.read_excel(f_rca, sheet_name='Sheet1'))
        df_rows2 = df2.shape[0]
        country_paticipant_quest_rca.to_excel(writer2, sheet_name='Sheet1', startrow=df_rows2 + 1, index=False, header=False)
        writer2.save()
        writer2.close()


time_end = time.time()
logger.info('The end,cost time:%d s', time_end - time_start)
