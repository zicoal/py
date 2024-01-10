from warnings import simplefilter
import matplotlib.pyplot as plt
import math
import os
import networkx as nx
import numpy as np
import pandas as pd
from random import *
import time
import logging
import xlsxwriter
from openpyxl import load_workbook
from xlsxwriter import Workbook

simplefilter(action='ignore', category=FutureWarning)

logger = logging.getLogger()
logger.setLevel(logging.INFO)
# 定义handler的输出格式
#logger to console
ch = logging.StreamHandler()
ch.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s")
#fh.setFormatter(formatter)
ch.setFormatter(formatter)
logger.addHandler(ch)

q_start_columm=2
rca_thresholds = [1, 1.1, 1.2, 1.3, 1.4, 1.5, 1.6,
                  1.7, 1.8, 1.9, 2]
rca_threshold=2.1




#equations = "score1-5"


#get the mapped score
#equations = "eq-nonlinear"
def getscore (x):
    b = 6
    c = 10
    return x**2 - b*x +c

#get the mapped score
equations = ["eq-linear","eq-nonlinear"]
#def getscore (x):
#    return x
#exit(0)



f1 = 'D:\\pydata\\data\\jonathan\\MCS_recoded.csv'
f_code= 'D:\\pydata\\data\\jonathan\\code.xlsx'

logger.info("loading data...")
time_start = time.time()
questions_original = ['q1', 'q2', 'q3', 'q4',
             'q5', 'q6', 'q7', 'q8',
             'q9', 'q10', 'q11', 'q12',
             'q13', 'q14', 'q15', 'q16']

questions = ['q1', 'q2', 'q4', 'q5',
              'q6', 'q7', 'q8','q9',
             'q10', 'q11', 'q12', 'q15']

df = pd.read_csv(f1)
# df = pd.read_excel(f1)
# df = df.loc[::, ['id','country', 'q1', 'q2', 'q3', 'q4',
#                 'q5', 'q6', 'q7', 'q8',
#                 'q9', 'q10', 'q11', 'q12',
#                 'q13', 'q14', 'q15', 'q16']]
# df_code= pd.read_excel(f_code)
# df_code = df_code.loc[::, ['q', 'answer', 'coded_value']]

time_end = time.time()
logger.info("data loaded...time cost:%d s'", time_end - time_start)
countries = {}
data = df.values.tolist()
countries = [data[i][1] for i in range(len(data))]
countries = sorted(list(set(countries)))
current_countries =countries

for equation in equations:
    for rca_threshold in rca_thresholds:
        f_rca= 'D:\\pydata\\data\\jonathan\\%s\\q%d\\rca_number_%.1f.csv' % (equation,len(questions), rca_threshold)
        f_weights = 'D:\\pydata\\data\\jonathan\\%s\\q%d\\weights_number_min_%.1f.csv' % (equation,len(questions), rca_threshold)
        logger.info("RCA:%.1f,Equation:%s" % (rca_threshold,equation))
        if os.path.isfile(f_rca)==False:
            country_paticipant_quest_rca = pd.DataFrame(columns=["country", 'id', 'Q', 'RCA'])
            country_quest_phi_weight_network = pd.DataFrame(columns=["country", 'QA', 'QB', 'weights'])

            country_paticipant_quest_rca.to_csv(f_rca, mode="w" ,index=False)

            country_quest_phi_weight_network.to_csv(f_weights, mode="w" , index=False)

        #dirty_words = {'Not sure', 'Dont\'t know','Refused'}



        k=0


        #current_countries = [countries[i] for i in range(k,len(countries))]
        #current_countries =["Zambia"]
        #current_countries =["Germany","Hong Kong","United States","Zambia"]

        #logger.info(current_countries)
        #logger.info(countries)

        #df_one_country=[]

        for y in countries:
            if (y in current_countries):
                k += 1
                time_end = time.time()
                country_paticipant_quest_rca = pd.DataFrame(columns=["country", 'id', 'Q', 'RCA'])
                country_quest_phi_weight_network = pd.DataFrame(columns=["country", 'QA', 'QB', 'weights'])
                df_one_country = df.loc[df['country'] == y]
                logger.info('RCA=:%.1f,%s,Country: %s(%d/%d),participants:%d, time:%d s',rca_threshold, equation,y, k,len(countries), len(df_one_country), time_end - time_start)
                data_one_country = df_one_country.values.tolist()
                df_one_country_values =  pd.DataFrame(columns=['id','q1', 'q2', 'q4', 'q5',
                  'q6', 'q7', 'q8','q9',
                 'q10', 'q11', 'q12', 'q15'])

                for i in range(len(data_one_country)):
                    q_values = [data_one_country[i][0]] #id
                    for j in range(len(questions)):
                            real_column =  questions_original.index(questions[j])
                            #print(questions[j],real_column+q_start_columm)
                            if equation == "eq-linear":
                                q_values.append(data_one_country[i][real_column+q_start_columm])
                            else:
                                q_values.append(getscore(data_one_country[i][real_column + q_start_columm]))
                    #exit(0)
                    df_one_country_values.loc[len(df_one_country_values)] = q_values
                paticipants = [data_one_country[i][0] for i in range(len(df_one_country))]
                paticipants = sorted(list(set(paticipants))) #去重
                #print(df_one_country_values)
                #exit(0)
                #计算每个问题的全局权重
                question_part={}
                df_temp = df_one_country_values.drop(['id'], axis=1)
        #        logger.info(df_temp)
                sum_quesitons=df_temp.sum()
                sum_all= sum(sum_quesitons)
                sum_participants=df_temp.sum(axis=1)

        #        logger.info(sum_all)
        #        logger.info(sum_quesitons)
        #        exit(0)
                question_participant = {}
                for i in range(len(questions)):
                    q = questions[i]
                    question_part[q] = sum_quesitons[i] / sum_all
                    question_participant[q]=[]
                    #logger.info(c,code_company.get(c))

                #logger.info(question_part)
                #exit(0)
                #计算单一question的RCA
                i=0
                for participant in paticipants:
                    #ignore invalid users : did not response at all
                    if (sum_participants[i] >0):
                        for j in range(len(questions)):
        #                logger.info('est')
        #                logger.info(df_one_country_values.loc[(df_one_country_values["id"]==participant)])
                        #logger.info(df_one_country_values.loc[(df_one_country_values["id"]==participant)][questions[j]])
        #                logger.info(df_one_country_values.loc[i,questions[j]])
        #                logger.info(sum_participants[i])
                            participant_question_ratio= df_one_country_values.loc[i,questions[j]] /sum_participants[i]
                            #participant_question_ratio= df_one_country_values.loc["id"==][questions[j]]/sum_participants[i]
                            #participant_question_global_ratio = df_one_country_values[i][j+1]/ df_one_country_values[i].sum(axis=0) #question_value
                            #binary computing RCA
            #                logger.info(participant_question_ratio)
            #                logger.info(question_part[questions[j]])
            #                exit(0)
            #                if df_one_country_values.loc[i,questions[j]]<=0:
                                #participant_question_ratio = 0
                            rca = 1 if participant_question_ratio/(question_part[questions[j]]) >= rca_threshold else 0
                            country_paticipant_quest_rca.loc[len(country_paticipant_quest_rca)] = [y, participant,questions[j], rca]

                           ###记录有效rca对应的人
                            if rca == 1:
                                list_participants=question_participant.get(questions[j])
            #                    logger.info(participant_question_ratio)
            #                    logger.info(question_part[questions[j]])
            #                    logger.info(participant)
            #                    exit(0)
                                list_participants.append(participant)
                                question_participant[questions[j]]=list_participants
                    i += 1
                #计算phi
                for i in range(len(questions)-1):
                    for j in range(i+1, len(questions)):
                        #两个集合的交集
                        set_participants= set(question_participant[questions[i]]) & set(question_participant[questions[j]])
                        if (len(set_participants)>0):
                            phi_ij = len(set_participants) / len(question_participant[questions[j]])
                            phi_ji = len(set_participants) / len(question_participant[questions[i]])
                            #@Todo: check  min or max
                            phi = phi_ij if phi_ij < phi_ji else phi_ji
                            country_quest_phi_weight_network.loc[len(country_quest_phi_weight_network)] = [y,questions[i], questions[j],phi]
                        #else:
                        #    country_quest_phi_weight_network.loc[len(country_quest_phi_weight_network)] = [y,questions[i], questions[j], 0]

                #write to file

                #time_end = time.time()
                country_quest_phi_weight_network.to_csv(f_weights, mode="a",  index=False, header=False)
    #            writer = pd.ExcelWriter(f_weights, engine='openpyxl',mode ='a', if_sheet_exists='overlay')
    #            book=load_workbook(f_weights)
                #@depressed: book and sheet setter for version under python 3.10
    #            writer.book = book
    #            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    #            df1=pd.DataFrame(pd.read_excel(f_weights,sheet_name='Sheet1'))
    #            df_rows = df1.shape[0]
    #            country_quest_phi_weight_network.to_excel(writer,sheet_name='Sheet1',startrow=df_rows+1,index=False,header=False)
    #            writer._save()
    #            writer.close()

                country_paticipant_quest_rca.to_csv(f_rca, mode="a",  index=False, header=False)
    #            writer2 = pd.CSVWriter(f_rca, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    #            book = load_workbook(f_rca)
    #            writer2.book = book
    #            writer2.sheets = dict((ws.title, ws) for ws in book.worksheets)
    #            df2 = pd.DataFrame(pd.read_excel(f_rca, sheet_name='Sheet1'))
    #            df_rows2 = df2.shape[0]
    #            country_paticipant_quest_rca.to_excel(writer2, sheet_name='Sheet1', startrow=df_rows2 + 1, index=False, header=False)
    #            writer2._save()
    #            writer2.close()


time_end = time.time()
logger.info('The end,cost time:%d s', time_end - time_start)
